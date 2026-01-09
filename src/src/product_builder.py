import os
import json

PROJECT_ROOT = os.getenv(
    "PROJECT_ROOT",
    os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
)

DRAFTS_DIR = os.path.join(PROJECT_ROOT, "data", "drafts", "templates")
PRODUCTS_DIR = os.path.join(PROJECT_ROOT, "data", "products")

os.makedirs(PRODUCTS_DIR, exist_ok=True)


def load_draft(draft_id: str):
    path = os.path.join(DRAFTS_DIR, f"{draft_id}.json")
    if not os.path.exists(path):
        raise FileNotFoundError(f"Draft not found: {path}")

    with open(path, "r") as f:
        return json.load(f)


# ---------------------------
# Product Builders
# ---------------------------

def build_excel_product(draft, product_path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tracker"

    ws["A1"] = draft["title"]
    ws["A3"] = "Date"
    ws["B3"] = "Task"
    ws["C3"] = "Status"
    ws["D3"] = "Notes"

    for i in range(4, 100):
        ws[f"A{i}"] = ""
        ws[f"B{i}"] = ""
        ws[f"C{i}"] = ""
        ws[f"D{i}"] = ""

    file_path = os.path.join(product_path, "tracker.xlsx")
    wb.save(file_path)

    return file_path


def build_printable_product(draft, product_path):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    file_path = os.path.join(product_path, "printable.pdf")
    c = canvas.Canvas(file_path, pagesize=A4)

    text = c.beginText(50, 800)
    text.setFont("Helvetica", 12)
    text.textLine(draft["title"])
    text.textLine("")
    text.textLine(draft["description"])
    text.textLine("")
    text.textLine("Features:")
    for f in draft.get("features", []):
        text.textLine(f"- {f}")

    c.drawText(text)
    c.save()

    return file_path


def build_canva_product(draft, product_path):
    file_path = os.path.join(product_path, "canva_template.txt")
    with open(file_path, "w") as f:
        f.write(f"CANVA TEMPLATE PLACEHOLDER\n\n{json.dumps(draft, indent=2)}")

    return file_path


def build_notion_product(draft, product_path):
    file_path = os.path.join(product_path, "notion_template.md")
    with open(file_path, "w") as f:
        f.write(f"# {draft['title']}\n\n")
        f.write(f"{draft['description']}\n\n")
        f.write("## Features\n")
        for feature in draft.get("features", []):
            f.write(f"- {feature}\n")

    return file_path


# ---------------------------
# Main Builder
# ---------------------------

def build_product_from_draft(draft_id: str):
    draft = load_draft(draft_id)

    product_id = str(uuid.uuid4())
    slug = draft["title"].lower().replace(" ", "_")[:40]
    product_path = os.path.join(PRODUCT_DIR, f"{slug}_{product_id}")
    os.makedirs(product_path, exist_ok=True)

    title = draft["title"].lower()

    if "excel" in title:
        artifact = build_excel_product(draft, product_path)
        product_type = "excel"

    elif "printable" in title or "planner" in title:
        artifact = build_printable_product(draft, product_path)
        product_type = "printable"

    elif "canva" in title:
        artifact = build_canva_product(draft, product_path)
        product_type = "canva"

    elif "notion" in title:
        artifact = build_notion_product(draft, product_path)
        product_type = "notion"

    else:
        artifact = build_notion_product(draft, product_path)
        product_type = "generic"

    metadata = {
        "product_id": product_id,
        "draft_id": draft_id,
        "title": draft["title"],
        "type": product_type,
        "artifact": artifact,
        "created_at": datetime.utcnow().isoformat()
    }

    with open(os.path.join(product_path, "product.json"), "w") as f:
        json.dump(metadata, f, indent=2)

    return metadata
